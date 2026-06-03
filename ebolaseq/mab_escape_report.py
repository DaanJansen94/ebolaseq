"""mAb escape epitope report for ebolaseq GP protein alignments."""
from __future__ import annotations

import json
import os
import re
from dataclasses import dataclass
from importlib import resources
from typing import Callable, Dict, List, Optional, Set, Tuple
from urllib.request import urlopen

from Bio import AlignIO, SeqIO
from Bio.Align import PairwiseAligner


@dataclass
class PositionEntry:
    pos: int
    domain: str
    region: str
    zaire_aa: str
    mAb114: bool
    REGN3470: bool
    REGN3471: bool
    REGN3479: bool
    ADI15878: bool
    ADI23774: bool
    contact: bool
    proven_escape: List[str]


@dataclass
class CellResult:
    aa: str
    status: str  # same, changed, gap, escape
    mutation_label: Optional[str] = None
    grantham: Optional[int] = None
    ref_aa: Optional[str] = None


@dataclass
class IsolateSummary:
    seq_id: str
    species: str
    n_changed: int = 0
    n_escape_match: int = 0
    cocktail_concern: bool = False
    mab114_concern: bool = False
    mAb114_hits: int = 0
    regn3470_hits: int = 0
    regn3471_hits: int = 0
    regn3479_hits: int = 0
    adi15878_hits: int = 0
    adi23774_hits: int = 0
    mbp134_concern: bool = False
    max_grantham: int = 0




def _reference_label(metadata: dict) -> str:
    return metadata.get("baseline_label", "Reference")


def _coordinate_mature_sequence(metadata: dict) -> str:
    """UniProt Q05320 mature GP for Q05320 position -> MSA column mapping."""
    if metadata.get("coordinate_mature_seq"):
        return metadata["coordinate_mature_seq"]
    mature_start = int(metadata.get("mature_gp_start", 33))
    acc = metadata.get("coordinate_uniprot", metadata.get("uniprot_accession", "Q05320"))
    url = f"https://rest.uniprot.org/uniprotkb/{acc}.fasta"
    with urlopen(url, timeout=30) as resp:
        lines = resp.read().decode("utf-8").splitlines()
    seq = "".join(line.strip() for line in lines if not line.startswith(">"))
    return seq[mature_start - 1 :]


def _baseline_mature_sequence(metadata: dict) -> str:
    """Makona 2014 mature GP (reference amino acids; file is already residues 33-676)."""
    if metadata.get("baseline_mature_seq"):
        return metadata["baseline_mature_seq"]
    fasta_name = metadata.get("baseline_fasta", "makona_gp_mature.fasta")
    try:
        pkg = resources.files("ebolaseq").joinpath("data", fasta_name)
        lines = pkg.read_text(encoding="utf-8").splitlines()
    except Exception:
        here = os.path.dirname(os.path.abspath(__file__))
        with open(os.path.join(here, "data", fasta_name), encoding="utf-8") as fh:
            lines = fh.read().splitlines()
    return "".join(line.strip() for line in lines if not line.startswith(">"))


def _q05320_mature_sequence(metadata: dict) -> str:
    return _coordinate_mature_sequence(metadata)

def load_watchlist(path: Optional[str] = None) -> Tuple[dict, List[PositionEntry]]:
    if path is None:
        try:
            pkg = resources.files("ebolaseq").joinpath("data/gp_epitope_watchlist.json")
            data = json.loads(pkg.read_text(encoding="utf-8"))
        except Exception:
            here = os.path.dirname(os.path.abspath(__file__))
            path = os.path.join(here, "data", "gp_epitope_watchlist.json")
            data = json.loads(open(path, encoding="utf-8").read())
    else:
        data = json.loads(open(path, encoding="utf-8").read())
    entries = [
        PositionEntry(
            pos=e["pos"], domain=e["domain"], region=e["region"], zaire_aa=e["zaire_aa"],
            mAb114=e["mAb114"], REGN3470=e["REGN3470"], REGN3471=e["REGN3471"],
            REGN3479=e["REGN3479"], ADI15878=e["ADI15878"], ADI23774=e["ADI23774"],
            contact=e["contact"], proven_escape=list(e.get("proven_escape", [])),
        )
        for e in data["positions"]
    ]
    return data.get("metadata", {}), entries


_GRANTHAM_CACHE: Optional[dict] = None


def _load_grantham_matrix() -> dict:
    global _GRANTHAM_CACHE
    if _GRANTHAM_CACHE is not None:
        return _GRANTHAM_CACHE
    try:
        pkg = resources.files("ebolaseq").joinpath("data/grantham_matrix.json")
        _GRANTHAM_CACHE = json.loads(pkg.read_text(encoding="utf-8"))
    except Exception:
        here = os.path.dirname(os.path.abspath(__file__))
        with open(os.path.join(here, "data", "grantham_matrix.json"), encoding="utf-8") as fh:
            _GRANTHAM_CACHE = json.load(fh)
    return _GRANTHAM_CACHE


def grantham_distance(ref_aa: str, alt_aa: str) -> Optional[int]:
    """Grantham distance (1974 Table 2); 5–215, higher = more dissimilar."""
    ref_aa, alt_aa = ref_aa.upper(), alt_aa.upper()
    if ref_aa == alt_aa:
        return 0
    m = _load_grantham_matrix()
    if ref_aa not in m or alt_aa not in m:
        return None
    v = m[ref_aa].get(alt_aa) or 0
    if v:
        return int(v)
    v = m[alt_aa].get(ref_aa) or 0
    return int(v) if v else None


def _ungapped(seq: str) -> str:
    return seq.replace("-", "").upper()


def _msa_col_for_ungapped_index(msa_row: str, uidx: int) -> Optional[int]:
    u = 0
    for col, aa in enumerate(msa_row):
        if aa != "-":
            if u == uidx:
                return col
            u += 1
    return None


def _is_makona_baseline_row(
    msa_row: str, baseline: str, pos_to_col: Dict[int, int], mature_start: int = 33
) -> bool:
    """True if MSA row matches packaged Makona mature GP at mapped Q05320 columns."""
    checked = 0
    for qpos, col in pos_to_col.items():
        idx = qpos - mature_start
        if idx < 0 or idx >= len(baseline):
            continue
        if col >= len(msa_row):
            continue
        aa = msa_row[col]
        if aa == "-":
            continue
        checked += 1
        if aa.upper() != baseline[idx].upper():
            return False
    return checked >= 10


def _build_makona_msa_row(
    pos_to_col: Dict[int, int],
    baseline: str,
    msa_length: int,
    mature_start: int = 33,
) -> str:
    """Place Makona mature GP into MSA columns using the Q05320 → column map."""
    row = ["-"] * msa_length
    for qpos, col in pos_to_col.items():
        idx = qpos - mature_start
        if 0 <= idx < len(baseline) and 0 <= col < msa_length:
            row[col] = baseline[idx]
    return "".join(row)


def _build_q05320_to_msa_col(q05320_mature: str, ref_msa_row: str) -> Dict[int, int]:
    ref_ungapped = _ungapped(ref_msa_row)
    aligner = PairwiseAligner(mode="global", match_score=2, mismatch_score=-1, open_gap_score=-2, extend_gap_score=-0.5)
    aln = aligner.align(q05320_mature, ref_ungapped)[0]
    # alignment: q05320 along target 0, ref along query 1
    q_map = {}
    qi = ri = 0
    for a, b in zip(aln[0], aln[1]):
        if a != "-" and b != "-":
            q_map[qi] = ri
            qi += 1
            ri += 1
        elif a != "-":
            qi += 1
        elif b != "-":
            ri += 1
    msa_map = {}
    mature_start = 33
    for qpos in range(mature_start, mature_start + len(q05320_mature)):
        mi = qpos - mature_start
        if mi not in q_map:
            continue
        col = _msa_col_for_ungapped_index(ref_msa_row, q_map[mi])
        if col is not None:
            msa_map[qpos] = col
    return msa_map


def _infer_species(seq_id: str) -> str:
    s = seq_id.lower()
    if "bundibugyo" in s or "/bdbv" in s or "bdbv" in s:
        return "Bundibugyo"
    if "zaire" in s or "zebov" in s or "ebola virus" in s:
        return "Zaire"
    if "sudan" in s:
        return "Sudan"
    if "reston" in s:
        return "Reston"
    if "tai" in s:
        return "Tai Forest"
    parts = seq_id.split("/")
    if len(parts) > 1:
        return parts[1]
    return "unknown"



def _accession(seq_id: str) -> str:
    token = seq_id.split()[0]
    return token.split("/")[0].split("|")[-1]


def _work_dir_from_aln(gp_aln_path: str) -> str:
    return os.path.abspath(os.path.join(os.path.dirname(gp_aln_path), "..", "..", ".."))


def _outgroup_accessions(work_dir: str) -> Set[str]:
    accessions: Set[str] = set()
    fasta_dir = os.path.join(work_dir, "FASTA")
    if not os.path.isdir(fasta_dir):
        return accessions
    for fn in os.listdir(fasta_dir):
        if not fn.startswith("outgroup_") or not fn.endswith((".fasta", ".fa")):
            continue
        for rec in SeqIO.parse(os.path.join(fasta_dir, fn), "fasta"):
            accessions.add(_accession(rec.id))
    return accessions


def _is_outgroup(seq_id: str, outgroup_accs: Set[str]) -> bool:
    if _accession(seq_id) in outgroup_accs:
        return True
    return "outgroup" in seq_id.lower()


def _yes_no(flag: bool) -> str:
    return "Yes" if flag else "No"


def _scope_filter_mab114(ent: PositionEntry) -> bool:
    return ent.mAb114


def _scope_filter_regn(ent: PositionEntry) -> bool:
    return ent.REGN3470 or ent.REGN3471 or ent.REGN3479


def _scope_filter_adi15878(ent: PositionEntry) -> bool:
    return ent.ADI15878


def _scope_filter_adi23774(ent: PositionEntry) -> bool:
    return ent.ADI23774


def _scope_filter_mbp134(ent: PositionEntry) -> bool:
    return ent.ADI15878 or ent.ADI23774


def _isolate_scope_stats(
    sid: str,
    matrix: dict,
    watchlist: List[PositionEntry],
    scope_filter: Callable[[PositionEntry], bool],
) -> dict:
    """Per-isolate counts within one therapy's watchlist footprint."""
    contact_positions = [e for e in watchlist if e.contact and scope_filter(e)]
    n_contact = len(contact_positions)
    n_same = sum(
        1 for ent in contact_positions if matrix[sid][ent.pos].status == "same"
    )
    pct = int(round(100 * n_same / n_contact)) if n_contact else 0
    n_changed = n_escape = 0
    max_g = 0
    for ent in watchlist:
        if not scope_filter(ent):
            continue
        c = matrix[sid][ent.pos]
        if c.status in ("changed", "escape") and (ent.contact or ent.proven_escape):
            n_changed += 1
        if (
            c.status == "escape"
            and c.mutation_label
            and c.mutation_label in ent.proven_escape
        ):
            n_escape += 1
        if c.grantham is not None:
            max_g = max(max_g, c.grantham)
    return {
        "n_same": n_same,
        "n_contact": n_contact,
        "pct": pct,
        "n_changed": n_changed,
        "n_escape": n_escape,
        "max_grantham": max_g,
    }


def _render_treatment_summary_table(
    matrix: dict,
    summaries: dict,
    watchlist: List[PositionEntry],
    table_id: str,
    title: str,
    scope_filter: Callable[[PositionEntry], bool],
    extra_columns: List[Tuple[str, str, str, Callable[[IsolateSummary, dict], Tuple[str, object]]]],
) -> str:
    """HTML sortable summary for one therapy. extra_columns: label, sort_key, sort_type, value_fn -> (display, sort_val)."""
    head_cells = [
        '<th class="sortable" data-sort-key="accession" data-sort-type="text">Accession</th>',
        '<th class="sortable" data-sort-key="species" data-sort-type="text">Species</th>',
        '<th class="sortable" data-sort-key="conserved" data-sort-type="number">Conserved contact sites</th>',
        '<th class="sortable" data-sort-key="epitope-changes" data-sort-type="number">Epitope changes</th>',
        '<th class="sortable" data-sort-key="escape" data-sort-type="number">Escape matches</th>',
        '<th class="sortable" data-sort-key="grantham" data-sort-type="number">Max Grantham</th>',
    ]
    for label, _sort_key, _sort_type, _value_fn in extra_columns:
        head_cells.append(
            f'<th class="sortable" data-sort-key="{_sort_key}" data-sort-type="{_sort_type}">{label}</th>'
        )
    rows = []
    n_cols = 6 + len(extra_columns)
    for sid in sorted(matrix.keys(), key=_accession):
        s = summaries[sid]
        st = _isolate_scope_stats(sid, matrix, watchlist, scope_filter)
        extra_attrs = ""
        extra_cells = ""
        for _label, sort_key, _sort_type, value_fn in extra_columns:
            display, sort_val = value_fn(s, st)
            extra_attrs += f' data-{sort_key}="{sort_val}"'
            extra_cells += f"<td>{display}</td>"
        rows.append(
            f'<tr data-accession="{_accession(sid)}" data-species="{s.species}"'
            f' data-conserved="{st["n_same"]}" data-conserved-pct="{st["pct"]}"'
            f' data-epitope-changes="{st["n_changed"]}" data-escape="{st["n_escape"]}"'
            f' data-grantham="{st["max_grantham"]}"{extra_attrs}>'
            f"<td>{_accession(sid)}</td><td>{s.species}</td>"
            f'<td>{st["n_same"]}/{st["n_contact"]} ({st["pct"]}%)</td>'
            f'<td>{st["n_changed"]}</td><td>{st["n_escape"]}</td><td>{st["max_grantham"]}</td>'
            f"{extra_cells}</tr>"
        )
    body = "".join(rows) if rows else f'<tr><td colspan="{n_cols}"><em>No isolates</em></td></tr>'
    return f"""<h3 class="summary-treatment-title">{title}</h3>
<table class="summary-table sortable-table" id="{table_id}"><thead><tr>
{"".join(head_cells)}
</tr></thead><tbody>{body}</tbody></table>"""


def _render_collapsible_section(
    section_id: str,
    heading: str,
    body_html: str,
    collapsed: bool = True,
    show_label: str = "Show",
    hide_label: str = "Hide",
) -> str:
    body_class = " treatment-collapsed" if collapsed else ""
    expanded = "false" if collapsed else "true"
    btn_label = show_label if collapsed else hide_label
    return (
        f'<div class="treatment-section" id="{section_id}">'
        f'<h2 class="treatment-heading">{heading} '
        f'<button type="button" class="toggle-treatment" aria-expanded="{expanded}" '
        f'aria-controls="{section_id}-body" data-show-label="{show_label}" '
        f'data-hide-label="{hide_label}">{btn_label}</button></h2>'
        f'<div class="treatment-body{body_class}" id="{section_id}-body">{body_html}</div>'
        "</div>"
    )


def _render_treatment_section(section_id: str, heading: str, body_html: str, collapsed: bool = True) -> str:
    """Collapsible block per therapeutic (Ebanga / Inmazeb / MBP134)."""
    return _render_collapsible_section(
        section_id, heading, body_html, collapsed, show_label="Show results", hide_label="Hide results"
    )


def _therapies_intro_html() -> str:
    return """<div class="card">
<p><strong>Ebanga (ansuvimab)</strong> &mdash; single monoclonal antibody <strong>mAb114</strong>.</p>
<p><strong>Inmazeb (REGN-EB3)</strong> &mdash; <strong>cocktail of three</strong> monoclonal antibodies:
<strong>mAb3470</strong> (atoltivimab), <strong>mAb3471</strong> (odesivimab), and <strong>mAb3479</strong> (maftivimab).</p>
<p><strong>MBP134 / MBP134AF</strong> &mdash; cocktail of <strong>ADI-15878</strong> and <strong>ADI-23774</strong>.</p>
</div>"""


def find_alignment_anchor_id(records, metadata: dict) -> str:
    """Pick an MSA row to map Q05320 coordinates (not the amino-acid comparison baseline)."""
    preferred = [
        r"kj660346",
        r"makona",
        r"kissidougou",
        r"gueckedou",
        r"/2014/",
        r"/2015/",
        r"/2016/",
        r"west.africa",
    ]
    fallback = [
        r"zaire.ebolavirus",
        r"zaire_ebolavirus",
        r"/zebov",
        r"NC_002549",
        r"mayinga",
        r"yambuku",
    ]
    for patterns in (preferred, fallback):
        for rec in records:
            h = rec.id.lower()
            desc = (rec.description or "").lower()
            if any(re.search(p, h) or re.search(p, desc) for p in patterns):
                return rec.id
    return records[0].id


def find_escape_baseline_reference_id(records, metadata: dict) -> str:
    return find_alignment_anchor_id(records, metadata)


def find_zaire_reference_id(records, metadata: dict) -> str:
    return find_escape_baseline_reference_id(records, metadata)


def _escape_label(ref_aa: str, pos: int, mut_aa: str, catalog: List[str]) -> Optional[str]:
    for lab in catalog:
        if len(lab) >= 4 and lab[0] == ref_aa and lab[-1] == mut_aa:
            try:
                p = int(lab[1:-1])
            except ValueError:
                continue
            if p == pos:
                return lab
    # generic
    if ref_aa != mut_aa and mut_aa not in ("-", "X"):
        return f"{ref_aa}{pos}{mut_aa}"
    return None


def score_alignment(gp_aln_path: str, watchlist: List[PositionEntry], metadata: dict, outgroup_accs: Optional[Set[str]] = None):
    records = list(AlignIO.read(gp_aln_path, "fasta"))
    if not records:
        raise ValueError("Empty GP protein alignment")
    anchor_id = find_alignment_anchor_id(records, metadata)
    anchor_row = None
    for rec in records:
        if rec.id == anchor_id:
            anchor_row = str(rec.seq)
            break
    if anchor_row is None:
        anchor_id = records[0].id
        anchor_row = str(records[0].seq)

    try:
        q05320_mature = _coordinate_mature_sequence(metadata)
    except Exception as exc:
        print(f"mab-escape-report: could not load Q05320 coordinate sequence ({exc}); using sparse watchlist map")
        mature_start = int(metadata.get("mature_gp_start", 33))
        max_pos = max(e.pos for e in watchlist)
        q05320_mature = ["X"] * (max_pos - mature_start + 1)
        for e in watchlist:
            q05320_mature[e.pos - mature_start] = e.zaire_aa
        q05320_mature = "".join(q05320_mature)

    pos_to_col = _build_q05320_to_msa_col(q05320_mature, anchor_row)
    missing = [e.pos for e in watchlist if e.pos not in pos_to_col]
    if missing:
        print(f"mab-escape-report: warning — could not map positions to MSA columns: {missing[:10]}{'...' if len(missing)>10 else ''}")

    matrix: Dict[str, Dict[int, CellResult]] = {}
    summaries: Dict[str, IsolateSummary] = {}

    if outgroup_accs is None:
        outgroup_accs = _outgroup_accessions(_work_dir_from_aln(gp_aln_path))
    excluded_outgroups: List[str] = []

    for rec in records:
        sid = rec.id
        if _is_outgroup(sid, outgroup_accs):
            excluded_outgroups.append(_accession(sid))
            continue
        row = str(rec.seq)
        species = _infer_species(sid + " " + (rec.description or ""))
        matrix[sid] = {}
        summ = IsolateSummary(seq_id=sid, species=species)
        for ent in watchlist:
            col = pos_to_col.get(ent.pos)
            if col is None:
                matrix[sid][ent.pos] = CellResult(aa="?", status="gap")
                continue
            aa = row[col] if col < len(row) else "-"
            if aa == "-":
                matrix[sid][ent.pos] = CellResult(aa="-", status="gap")
                if ent.contact:
                    summ.n_changed += 1
                continue
            aa = aa.upper()
            ref_aa = ent.zaire_aa.upper()
            if aa == ref_aa:
                matrix[sid][ent.pos] = CellResult(aa=aa, status="same")
                continue
            label = _escape_label(ref_aa, ent.pos, aa, ent.proven_escape)
            status = "escape" if label in ent.proven_escape else "changed"
            g = grantham_distance(ref_aa, aa)
            matrix[sid][ent.pos] = CellResult(aa=aa, status=status, mutation_label=label, grantham=g, ref_aa=ref_aa)
            if ent.contact or ent.proven_escape:
                summ.n_changed += 1
            if label in ent.proven_escape:
                summ.n_escape_match += 1
            if ent.mAb114 and (ent.contact or ent.proven_escape):
                summ.mAb114_hits += 1
            if ent.REGN3470 and (ent.contact or ent.proven_escape):
                summ.regn3470_hits += 1
            if ent.REGN3471 and (ent.contact or ent.proven_escape):
                summ.regn3471_hits += 1
            if ent.REGN3479 and (ent.contact or ent.proven_escape):
                summ.regn3479_hits += 1
            if ent.ADI15878 and (ent.contact or ent.proven_escape):
                summ.adi15878_hits += 1
            if ent.ADI23774 and (ent.contact or ent.proven_escape):
                summ.adi23774_hits += 1

        for _ent in watchlist:
            _c = matrix[sid].get(_ent.pos)
            if _c and _c.grantham is not None:
                summ.max_grantham = max(summ.max_grantham, _c.grantham)
        summ.mab114_concern = summ.mAb114_hits > 0
        components_hit = sum([
            summ.regn3470_hits > 0,
            summ.regn3471_hits > 0,
            summ.regn3479_hits > 0,
        ])
        summ.cocktail_concern = components_hit >= 2
        mbp_components_hit = sum([
            summ.adi15878_hits > 0,
            summ.adi23774_hits > 0,
        ])
        summ.mbp134_concern = mbp_components_hit >= 2
        summaries[sid] = summ

    if excluded_outgroups:
        print(
            "mab-escape-report: excluded phylogeny outgroup(s): "
            + ", ".join(sorted(set(excluded_outgroups)))
        )
    return anchor_id, pos_to_col, matrix, summaries, excluded_outgroups



def _has_change_in_scope(sid, matrix, watchlist, position_filter):
    for ent in watchlist:
        if not position_filter(ent):
            continue
        if matrix[sid][ent.pos].status in ("changed", "escape"):
            return True
    return False


def _changed_isolates(matrix, summaries, _anchor_id, watchlist, position_filter=None):
    pf = position_filter or (lambda e: True)
    result = []
    for sid in matrix:
        if summaries[sid].n_changed == 0 and summaries[sid].n_escape_match == 0:
            continue
        if _has_change_in_scope(sid, matrix, watchlist, pf):
            result.append(sid)
    return sorted(result, key=_accession)


def _canonical_isolate_order(matrix: dict, _anchor_id: str) -> List[str]:
    return sorted(matrix.keys(), key=_accession)


def _xlsx_sheet(ws, header: List[str], rows: List[List]) -> None:
    ws.append(header)
    for row in rows:
        ws.append(row)


def write_r_workbook(
    output_dir: str,
    gp_aln_path: str,
    watchlist: List[PositionEntry],
    ref_id: str,
    matrix: dict,
    summaries: dict,
    metadata: Optional[dict] = None,
) -> str:
    """One Excel file (tabs) with tables needed for R figures."""
    from openpyxl import Workbook

    work_dir = _work_dir_from_aln(gp_aln_path)
    consensus_hints = _collect_user_consensus_ids(work_dir)
    nonref = sorted(matrix.keys(), key=_accession)
    n_consensus = sum(1 for sid in nonref if _sequence_source(sid, consensus_hints) == "consensus")
    n_genbank = len(nonref) - n_consensus
    if consensus_hints:
        in_aln = {sid for sid in nonref if _sequence_source(sid, consensus_hints) == "consensus"}
        missing = []
        for hint in sorted(consensus_hints, key=str.lower):
            if hint in matrix:
                continue
            if any(_is_user_consensus(sid, {hint}) for sid in matrix):
                continue
            if _accession(hint) in {_accession(s) for s in in_aln}:
                continue
            missing.append(hint)
        if missing:
            print(
                "mab-escape-report: warning — %d user consensus ID(s) not in GP alignment "
                "(CDS/GP extraction may have dropped them): %s"
                % (len(missing), ", ".join(missing[:5]) + ("..." if len(missing) > 5 else ""))
            )

    cell_header = [
        "sequence_id",
        "accession",
        "source",
        "species",
        "pos",
        "region",
        "ref_aa",
        "aa",
        "status",
        "mutation_label",
        "grantham",
        "contact",
        "mAb114",
        "REGN3471",
    ]
    cell_rows = []
    cell_rows_consensus = []
    cell_rows_genbank = []
    for sid in nonref:
        summ = summaries.get(sid)
        species = summ.species if summ else _infer_species(sid)
        source = _sequence_source(sid, consensus_hints)
        for ent in watchlist:
            c = matrix[sid][ent.pos]
            row = [
                sid,
                _accession(sid),
                source,
                species,
                ent.pos,
                ent.region,
                ent.zaire_aa,
                c.aa,
                c.status,
                c.mutation_label or "",
                c.grantham if c.grantham is not None else "",
                ent.contact,
                ent.mAb114,
                ent.REGN3471,
            ]
            cell_rows.append(row)
            if source == "consensus":
                cell_rows_consensus.append(row)
            else:
                cell_rows_genbank.append(row)

    iso_mab114_header = [
        "sequence_id",
        "accession",
        "source",
        "species",
        "n_conserved_contact",
        "n_contact_sites",
        "pct_conserved_contact",
        "epitope_changes",
        "escape_matches",
        "max_grantham",
        "mAb114_concern",
    ]
    iso_regn_header = [
        "sequence_id",
        "accession",
        "source",
        "species",
        "n_conserved_contact",
        "n_contact_sites",
        "pct_conserved_contact",
        "epitope_changes",
        "escape_matches",
        "max_grantham",
        "mAb3470_hits",
        "mAb3471_hits",
        "mAb3479_hits",
        "REGN_EB3_cocktail_concern",
    ]
    iso_mab114_rows = []
    iso_regn_rows = []
    iso_mab114_rows_consensus = []
    iso_regn_rows_consensus = []
    iso_mbp134_rows_consensus = []
    iso_mbp134_header = [
        "sequence_id",
        "accession",
        "source",
        "species",
        "n_conserved_contact",
        "n_contact_sites",
        "pct_conserved_contact",
        "epitope_changes",
        "escape_matches",
        "max_grantham",
        "ADI-15878_hits",
        "ADI-23774_hits",
        "MBP134_cocktail_concern",
    ]
    iso_mbp134_rows = []
    for sid in nonref:
        summ = summaries[sid]
        source = _sequence_source(sid, consensus_hints)
        st114 = _isolate_scope_stats(sid, matrix, watchlist, _scope_filter_mab114)
        stregn = _isolate_scope_stats(sid, matrix, watchlist, _scope_filter_regn)
        stmbp = _isolate_scope_stats(sid, matrix, watchlist, _scope_filter_mbp134)
        pct114 = round(100 * st114["n_same"] / st114["n_contact"], 1) if st114["n_contact"] else ""
        pcregn = round(100 * stregn["n_same"] / stregn["n_contact"], 1) if stregn["n_contact"] else ""
        pctmbp = round(100 * stmbp["n_same"] / stmbp["n_contact"], 1) if stmbp["n_contact"] else ""
        row114 = [
            sid,
            _accession(sid),
            source,
            summ.species,
            st114["n_same"],
            st114["n_contact"],
            pct114,
            st114["n_changed"],
            st114["n_escape"],
            st114["max_grantham"],
            summ.mab114_concern,
        ]
        rowregn = [
            sid,
            _accession(sid),
            source,
            summ.species,
            stregn["n_same"],
            stregn["n_contact"],
            pcregn,
            stregn["n_changed"],
            stregn["n_escape"],
            stregn["max_grantham"],
            summ.regn3470_hits,
            summ.regn3471_hits,
            summ.regn3479_hits,
            summ.cocktail_concern,
        ]
        rowmbp = [
            sid,
            _accession(sid),
            source,
            summ.species,
            stmbp["n_same"],
            stmbp["n_contact"],
            pctmbp,
            stmbp["n_changed"],
            stmbp["n_escape"],
            stmbp["max_grantham"],
            summ.adi15878_hits,
            summ.adi23774_hits,
            summ.mbp134_concern,
        ]
        iso_mab114_rows.append(row114)
        iso_regn_rows.append(rowregn)
        iso_mbp134_rows.append(rowmbp)
        if source == "consensus":
            iso_mab114_rows_consensus.append(row114)
            iso_regn_rows_consensus.append(rowregn)
            iso_mbp134_rows_consensus.append(rowmbp)

    catalog = _collect_proven_escape_catalog(watchlist, metadata)
    esc_header = ["mutation", "pos", "in_dataset", "n_isolates"]
    esc_rows = []
    for item in catalog:
        hits = _isolates_with_escape(matrix, item["pos"], item["label"], nonref)
        esc_rows.append([item["label"], item["pos"], bool(hits), len(hits)])

    path = os.path.join(output_dir, "mab_escape_data.xlsx")
    wb = Workbook()
    ws_cells = wb.active
    ws_cells.title = "epitope_cells"
    _xlsx_sheet(ws_cells, cell_header, cell_rows)
    ws_mab114 = wb.create_sheet("isolates_mab114")
    _xlsx_sheet(ws_mab114, iso_mab114_header, iso_mab114_rows)
    ws_regn = wb.create_sheet("isolates_regn")
    _xlsx_sheet(ws_regn, iso_regn_header, iso_regn_rows)
    ws_mbp = wb.create_sheet("isolates_MBP134")
    _xlsx_sheet(ws_mbp, iso_mbp134_header, iso_mbp134_rows)
    ws_esc = wb.create_sheet("escape_catalog")
    _xlsx_sheet(ws_esc, esc_header, esc_rows)
    if cell_rows_consensus:
        ws = wb.create_sheet("epitope_cells_consensus")
        _xlsx_sheet(ws, cell_header, cell_rows_consensus)
        ws = wb.create_sheet("isolates_mab114_consensus")
        _xlsx_sheet(ws, iso_mab114_header, iso_mab114_rows_consensus)
        ws = wb.create_sheet("isolates_regn_consensus")
        _xlsx_sheet(ws, iso_regn_header, iso_regn_rows_consensus)
        ws = wb.create_sheet("isolates_MBP134_consensus")
        _xlsx_sheet(ws, iso_mbp134_header, iso_mbp134_rows_consensus)
    if cell_rows_genbank:
        ws = wb.create_sheet("epitope_cells_genbank")
        _xlsx_sheet(ws, cell_header, cell_rows_genbank)
    wb.save(path)
    print(
        "mab-escape-report: xlsx isolates — %d genbank, %d consensus (in GP alignment)"
        % (n_genbank, n_consensus)
    )
    return path




def _cell_class(c: CellResult) -> str:
    return {"same": "same", "changed": "changed", "escape": "escape", "gap": "gap"}.get(c.status, "")


def _format_cell_html(c: CellResult) -> str:
    if c.status in ("changed", "escape"):
        label = c.mutation_label if c.mutation_label else c.aa
        gs = f' <span class="gscore">({c.grantham})</span>' if c.grantham is not None else ""
        return f'<span class="mut">{label}</span>{gs}'
    return c.aa


def _grantham_legend_html() -> str:
    return """<div class="card">
<p>In epitope tables, changed residues are shown as <span class="mut">MUTATION</span><span class="gscore"> (score)</span>
&mdash; for example <span class="mut">T144M</span><span class="gscore"> (110)</span>. The number in brackets is the
<strong>Grantham amino acid difference index</strong> (Grantham, 1974) versus the reference baseline amino acid at that
position. Scores run from <strong>0</strong> (identical) to about <strong>215</strong> (most dissimilar single-residue
substitutions in the matrix). Higher scores mean a more radical biochemical change (size, charge, polarity).</p>
<ul class="grantham-scale">
<li><strong>0</strong> &mdash; identical to reference</li>
<li><strong>1&ndash;50</strong> &mdash; conservative substitution (similar properties)</li>
<li><strong>51&ndash;100</strong> &mdash; moderate difference</li>
<li><strong>101&ndash;150</strong> &mdash; substantial difference</li>
<li><strong>151&ndash;215</strong> &mdash; radical substitution (very dissimilar properties)</li>
</ul>
</div>"""




def _escape_linked_mabs(label: str, watchlist: List[PositionEntry], escape_mab_map: dict) -> List[str]:
    """Map escape mutation to mAb names; escape_mab_map overrides shared epitope positions."""
    key_map = {
        "mAb114": "mAb114",
        "REGN3470": "mAb3470",
        "REGN3471": "mAb3471",
        "REGN3479": "mAb3479",
        "ADI15878": "ADI-15878",
        "ADI23774": "ADI-23774",
    }
    if label in escape_mab_map:
        return [key_map[k] for k in escape_mab_map[label] if k in key_map]
    mabs: List[str] = []
    for e in watchlist:
        if label not in e.proven_escape:
            continue
        if e.mAb114 and "mAb114" not in mabs:
            mabs.append("mAb114")
        if e.REGN3470 and "mAb3470" not in mabs:
            mabs.append("mAb3470")
        if e.REGN3471 and "mAb3471" not in mabs:
            mabs.append("mAb3471")
        if e.REGN3479 and "mAb3479" not in mabs:
            mabs.append("mAb3479")
        if e.ADI15878 and "ADI-15878" not in mabs:
            mabs.append("ADI-15878")
        if e.ADI23774 and "ADI-23774" not in mabs:
            mabs.append("ADI-23774")
    return mabs


def _collect_proven_escape_catalog(watchlist: List[PositionEntry], metadata: Optional[dict] = None) -> List[dict]:
    """Unique published escape mutations from watchlist with linked antibodies."""
    metadata = metadata or {}
    escape_mab_map = metadata.get("escape_mab_map", {})
    seen: Set[str] = set()
    items: List[dict] = []
    for ent in watchlist:
        for lab in ent.proven_escape:
            if lab in seen:
                continue
            seen.add(lab)
            mabs = _escape_linked_mabs(lab, watchlist, escape_mab_map)
            products: List[str] = []
            if "mAb114" in mabs:
                products.append("Ebanga")
            if any(m in mabs for m in ("mAb3470", "mAb3471", "mAb3479")):
                products.append("Inmazeb")
            if any(m in mabs for m in ("ADI-15878", "ADI-23774")):
                products.append("MBP134")
            items.append({
                "label": lab,
                "pos": ent.pos,
                "zaire_aa": ent.zaire_aa,
                "region": ent.region,
                "mabs": mabs,
                "products": products,
            })
    return sorted(items, key=lambda x: (x["pos"], x["label"]))


def _isolates_with_escape(matrix: dict, pos: int, label: str, isolates: List[str]) -> List[str]:
    found = []
    for sid in isolates:
        c = matrix[sid].get(pos)
        if c and c.status == "escape" and c.mutation_label == label:
            found.append(_accession(sid))
    return sorted(found)


def _proven_escape_section(
    matrix: dict, ref_id: str, watchlist: List[PositionEntry], metadata: Optional[dict] = None
) -> str:
    catalog = _collect_proven_escape_catalog(watchlist, metadata)
    if not catalog:
        return ""
    isolates = sorted(matrix.keys(), key=_accession)
    rows = []
    n_found_any = 0
    for item in catalog:
        hits = _isolates_with_escape(matrix, item["pos"], item["label"], isolates)
        if hits:
            n_found_any += 1
        in_data = f'<strong>Yes</strong> ({len(hits)} isolate(s))' if hits else "No"
        acc_cell = ", ".join(hits) if hits else "&mdash;"
        mab_str = ", ".join(item["mabs"])
        prod_str = ", ".join(item.get("products", [])) or "&mdash;"
        row_class = "escape" if hits else ""
        rows.append(
            f"<tr><td><strong>{item['label']}</strong></td><td>{item['pos']}</td>"
            f"<td>{item['zaire_aa']}&rarr;{item['label'][-1]}</td><td>{mab_str}</td>"
            f"<td>{prod_str}</td>"
            f'<td>{item["region"]}</td><td class="{row_class}">{in_data}</td>'
            f"<td>{acc_cell}</td></tr>"
        )
    ref_label = _reference_label(metadata or {})
    return f"""<p>Published in vitro escape variants from the watchlist, checked against all analyzed isolates (vs {ref_label}). This is not an exhaustive list of all possible escapes.</p>
<p><strong>Detected in this run:</strong> {n_found_any}/{len(catalog)} catalogued mutation(s).</p>
<table class="proven-escape-table"><thead><tr>
<th>Mutation</th><th>Pos</th><th>Change</th><th>Linked mAb</th><th>Product(s)</th><th>Region</th><th>In dataset?</th><th>Accessions</th>
</tr></thead><tbody>{"".join(rows)}</tbody></table>"""



_CONSENSUS_FILE_TO_SPECIES = {
    "consensus_zaire": "Zaire_ebolavirus",
    "consensus_sudan": "Sudan_ebolavirus",
    "consensus_reston": "Reston_ebolavirus",
    "consensus_bundibugyo": "Bundibugyo_ebolavirus",
    "consensus_tai_forest": "Tai_Forest_ebolavirus",
    "consensus_tai": "Tai_Forest_ebolavirus",
}
_KNOWN_PROCESSING_SPECIES = set(_CONSENSUS_FILE_TO_SPECIES.values())


def _normalize_consensus_seq_id(raw_id: str, species: str, index: int) -> str:
    """Match ID normalization in ebolaseq.run_protein_pipeline combined FASTA build."""
    parts = raw_id.split("/")
    if len(parts) >= 2 and parts[1] in _KNOWN_PROCESSING_SPECIES:
        return raw_id
    base_id = (raw_id.split()[0] if raw_id else "consensus").replace("/", "_")
    return "%s_%d/%s/consensus" % (base_id, index, species)


def _collect_user_consensus_ids(work_dir: str) -> Set[str]:
    """IDs from FASTA/consensus_*.fasta (normalized + raw) for cohort tagging."""
    ids: Set[str] = set()
    fasta_dir = os.path.join(work_dir, "FASTA")
    if not os.path.isdir(fasta_dir):
        return ids
    for fname in sorted(os.listdir(fasta_dir)):
        if not fname.startswith("consensus_"):
            continue
        if not (fname.endswith(".fasta") or fname.endswith(".fa")):
            continue
        base = fname.replace(".fasta", "").replace(".fa", "")
        species = _CONSENSUS_FILE_TO_SPECIES.get(base)
        path = os.path.join(fasta_dir, fname)
        idx = 0
        for rec in SeqIO.parse(path, "fasta"):
            raw = rec.id
            ids.add(raw)
            if species:
                norm = _normalize_consensus_seq_id(raw, species, idx)
                ids.add(norm)
                idx += 1
    combined = os.path.join(fasta_dir, "Ebola_Combined.fasta")
    if os.path.isfile(combined):
        for rec in SeqIO.parse(combined, "fasta"):
            if _is_user_consensus(rec.id):
                ids.add(rec.id)
    return ids


def _is_user_consensus(seq_id: str, consensus_hints: Optional[Set[str]] = None) -> bool:
    """Sequences from ebolaseq --c_z, --c_b, etc."""
    if not seq_id:
        return False
    if seq_id in (consensus_hints or ()):
        return True
    low = seq_id.lower()
    if "/consensus" in low:
        return True
    if "|" in seq_id and "/consensus" in seq_id.split("|", 1)[-1].lower():
        return True
    if consensus_hints:
        tail = seq_id.split("|")[-1] if "|" in seq_id else seq_id
        if tail in consensus_hints:
            return True
    return False


def _sequence_source(seq_id: str, consensus_hints: Set[str]) -> str:
    return "consensus" if _is_user_consensus(seq_id, consensus_hints) else "genbank"


def _cohort_subset(
    matrix: dict, summaries: dict, cohort: str, consensus_hints: Optional[Set[str]] = None
) -> Tuple[dict, dict]:
    hints = consensus_hints or set()
    if cohort == "consensus":
        keys = [k for k in matrix if _is_user_consensus(k, hints)]
    else:
        keys = [k for k in matrix if not _is_user_consensus(k, hints)]
    return {k: matrix[k] for k in keys}, {k: summaries[k] for k in keys if k in summaries}


def _render_report_panel(
    matrix: dict,
    summaries: dict,
    ref_id: str,
    watchlist: List[PositionEntry],
    panel_prefix: str = "",
    metadata: Optional[dict] = None,
) -> str:
    if not matrix:
        return "<p><em>No sequences in this set.</em></p>"
    metadata = metadata or {}
    ref_label = _reference_label(metadata)
    ref_col = f"{ref_label} aa"
    all_changed = _changed_isolates(matrix, summaries, ref_id, watchlist)
    mab114_table_id = f"{panel_prefix}-summary-mab114" if panel_prefix else "summary-mab114"
    regn_table_id = f"{panel_prefix}-summary-regn" if panel_prefix else "summary-regn"
    mbp134_table_id = f"{panel_prefix}-summary-mbp134" if panel_prefix else "summary-mbp134"
    mab114_summary = _render_treatment_summary_table(
        matrix,
        summaries,
        watchlist,
        mab114_table_id,
        "Ebanga (mAb114 / ansuvimab)",
        _scope_filter_mab114,
        [
            (
                "Concern",
                "concern",
                "number",
                lambda s, _st: (_yes_no(s.mab114_concern), 1 if s.mab114_concern else 0),
            ),
        ],
    )
    regn_summary = _render_treatment_summary_table(
        matrix,
        summaries,
        watchlist,
        regn_table_id,
        "Inmazeb (REGN-EB3: mAb3470 + mAb3471 + mAb3479)",
        _scope_filter_regn,
        [
            (
                "mAb3470 hits",
                "m3470",
                "number",
                lambda s, _st: (s.regn3470_hits, s.regn3470_hits),
            ),
            (
                "mAb3471 hits",
                "m3471",
                "number",
                lambda s, _st: (s.regn3471_hits, s.regn3471_hits),
            ),
            (
                "mAb3479 hits",
                "m3479",
                "number",
                lambda s, _st: (s.regn3479_hits, s.regn3479_hits),
            ),
            (
                "Cocktail concern (&ge;2 components)",
                "cocktail-concern",
                "number",
                lambda s, _st: (_yes_no(s.cocktail_concern), 1 if s.cocktail_concern else 0),
            ),
        ],
    )
    mbp134_summary = _render_treatment_summary_table(
        matrix,
        summaries,
        watchlist,
        mbp134_table_id,
        "MBP134 (ADI-15878 + ADI-23774)",
        _scope_filter_mbp134,
        [
            (
                "ADI-15878 hits",
                "a15878",
                "number",
                lambda s, _st: (s.adi15878_hits, s.adi15878_hits),
            ),
            (
                "ADI-23774 hits",
                "a23774",
                "number",
                lambda s, _st: (s.adi23774_hits, s.adi23774_hits),
            ),
            (
                "Cocktail concern (both components)",
                "mbp-concern",
                "number",
                lambda s, _st: (_yes_no(s.mbp134_concern), 1 if s.mbp134_concern else 0),
            ),
        ],
    )

    def analysis_isolates():
        return sorted(matrix.keys(), key=_accession)

    def antibody_table(title: str, position_filter):
        positions = [ent for ent in watchlist if position_filter(ent)]
        if not positions:
            return f"<h2>{title}</h2><p><em>No positions in watchlist for this antibody.</em></p>"
        slug = re.sub(r"[^a-z0-9]+", "-", title.lower()).strip("-")
        table_id = f"{panel_prefix}-{slug}" if panel_prefix else slug
        isolates = _canonical_isolate_order(matrix, ref_id)
        changed_rows = 0
        rows = []
        for row_num, ent in enumerate(positions, start=1):
            row_max = row_has_change = 0
            for sid in isolates:
                c = matrix[sid][ent.pos]
                if c.status in ("changed", "escape"):
                    row_has_change = 1
                if c.grantham is not None:
                    row_max = max(row_max, c.grantham)
            if row_has_change:
                changed_rows += 1
            tds = [
                f"<td>{row_num}</td>",
                f"<td>{ent.pos}</td>",
                f"<td>{ent.domain}</td>",
                f"<td>{ent.zaire_aa}</td>",
            ]
            for sid in isolates:
                c = matrix[sid][ent.pos]
                g = c.grantham if c.grantham is not None else 0
                tds.append(f'<td class="{_cell_class(c)}" data-grantham="{g}">{_format_cell_html(c)}</td>')
            conserved_attr = "" if row_has_change else ' class="conserved-row conserved-hidden"'
            rows.append(
                f'<tr data-row-max-grantham="{row_max}" data-pos="{ent.pos}"{conserved_attr}>'
                + "".join(tds)
                + "</tr>"
            )
        headers = "".join(f"<th>{_accession(sid)}</th>" for sid in isolates)
        if changed_rows:
            change_line = (
                f"<p><strong>Epitope changes vs {ref_label}:</strong> "
                f"{changed_rows} position(s) with &ge;1 change.</p>"
            )
        else:
            change_line = (
                f"<p><strong>Epitope changes vs {ref_label}:</strong> "
                f"no changes at these watchlist sites.</p>"
            )
        return f"""<h2>{title}</h2>
{change_line}
<button type="button" class="toggle-conserved" data-table="{table_id}">Show conserved rows</button>
<table class="epitope-table" id="{table_id}"><thead><tr><th>#</th><th>Pos</th><th>Domain</th><th>{ref_col}</th>{headers}</tr></thead>
<tbody>{"".join(rows)}</tbody></table>"""

    no_changes = ""
    if not all_changed:
        no_changes = f"<p><strong>All epitope positions match {ref_label}</strong> in this sequence set.</p>"

    prefix = f"{panel_prefix}-" if panel_prefix else ""
    ebanga_body = "\n".join(
        [mab114_summary, antibody_table("mAb114 (Ebanga / ansuvimab)", lambda e: e.mAb114)]
    )
    inmazeb_body = "\n".join(
        [
            regn_summary,
            antibody_table("mAb3470 — atoltivimab (Inmazeb)", lambda e: e.REGN3470),
            antibody_table("mAb3471 — odesivimab (Inmazeb)", lambda e: e.REGN3471),
            antibody_table("mAb3479 — maftivimab (Inmazeb)", lambda e: e.REGN3479),
        ]
    )
    mbp134_body = "\n".join(
        [
            mbp134_summary,
            antibody_table("ADI-15878 (MBP134 component 1)", lambda e: e.ADI15878),
            antibody_table("ADI-23774 (MBP134 component 2)", lambda e: e.ADI23774),
        ]
    )
    parts = [
        no_changes,
        _render_treatment_section(
            f"{prefix}treatment-ebanga",
            "Ebanga (mAb114 / ansuvimab)",
            ebanga_body,
        ),
        _render_treatment_section(
            f"{prefix}treatment-inmazeb",
            "Inmazeb (REGN-EB3)",
            inmazeb_body,
        ),
        _render_treatment_section(
            f"{prefix}treatment-mbp134",
            "MBP134 (ADI-15878 + ADI-23774)",
            mbp134_body,
        ),
    ]
    escape_catalog_body = _proven_escape_section(matrix, ref_id, watchlist, metadata)
    if escape_catalog_body:
        parts.append(
            _render_treatment_section(
                f"{prefix}treatment-escape-catalog",
                "Catalogued escape mutations (literature)",
                escape_catalog_body,
            )
        )
    return "\n".join(parts)


def write_html(
    path: str,
    gp_aln_path: str,
    metadata: dict,
    watchlist: List[PositionEntry],
    ref_id: str,
    matrix: dict,
    summaries: dict,
    excluded_outgroups: List[str],
):
    work_dir = _work_dir_from_aln(gp_aln_path)
    consensus_hints = _collect_user_consensus_ids(work_dir)
    m_dl, s_dl = _cohort_subset(matrix, summaries, "downloaded", consensus_hints)
    m_cs, s_cs = _cohort_subset(matrix, summaries, "consensus", consensus_hints)
    n_dl = len(m_dl)
    n_cs = len(m_cs)
    has_tabs = n_cs > 0

    ref_label = _reference_label(metadata)
    print(
        "mab-escape-report: comparison baseline %s (%s); MSA coordinate anchor %s"
        % (ref_label, metadata.get("reference_accession", ""), _accession(ref_id))
    )
    panel_downloaded = _render_report_panel(
        m_dl, s_dl, ref_id, watchlist, panel_prefix="downloaded", metadata=metadata
    )
    panel_consensus = (
        _render_report_panel(
            m_cs, s_cs, ref_id, watchlist, panel_prefix="consensus", metadata=metadata
        )
        if has_tabs
        else ""
    )

    tabs_html = ""
    panels_html = ""
    if has_tabs:
        tabs_html = """<div class="tabs" role="tablist">
<button type="button" class="tab-btn active" data-panel="panel-downloaded" role="tab">Downloaded (GenBank)</button>
<button type="button" class="tab-btn" data-panel="panel-consensus" role="tab">Added sequences (consensus)</button>
</div>"""
        panels_html = f"""<div id="panel-downloaded" class="tab-panel active" role="tabpanel">{panel_downloaded}</div>
<div id="panel-consensus" class="tab-panel" role="tabpanel">{panel_consensus}</div>"""
    else:
        panels_html = f'<div class="tab-panel active">{panel_downloaded}</div>'

    html = f"""<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8">
<title>GP mAb escape report</title>
<style>
:root {{ --bg:#f8f9fb; --card:#fff; --accent:#1a5fb4; --border:#d0d7de; }}
body {{ font-family: system-ui, -apple-system, Segoe UI, sans-serif; margin: 0; background: var(--bg); color: #222; line-height: 1.5; }}
.wrap {{ max-width: 1280px; margin: 0 auto; padding: 1.25rem 1.5rem 2rem; }}
h1 {{ font-size: 1.5rem; margin: 0 0 0.5rem; }}
h2 {{ font-size: 1.15rem; margin-top: 1.75rem; border-bottom: 2px solid var(--accent); padding-bottom: 0.25rem; }}
.card {{ background: var(--card); border: 1px solid var(--border); border-radius: 8px; padding: 1rem 1.15rem; margin: 1rem 0; box-shadow: 0 1px 2px rgba(0,0,0,.04); }}
.card h3 {{ margin: 0 0 0.5rem; font-size: 1rem; color: var(--accent); }}
.grantham-scale {{ margin: 0.35rem 0 0.65rem 1.15rem; font-size: 0.9rem; padding: 0; }}
.grantham-scale li {{ margin: 0.2rem 0; }}
.tabs {{ display: flex; flex-wrap: wrap; gap: 0.35rem; margin: 1.25rem 0 0; }}
.tab-btn {{ padding: 0.45rem 1rem; border: 1px solid var(--border); background: #fff; cursor: pointer; border-radius: 6px; font-size: 0.9rem; }}
.tab-btn.active {{ background: var(--accent); color: #fff; border-color: var(--accent); }}
.tab-panel {{ display: none; margin-top: 0.5rem; }}
.tab-panel.active {{ display: block; }}
.controls {{ display: flex; flex-wrap: wrap; gap: 1rem 2rem; align-items: center; background: #eef2f7; padding: 0.75rem 1rem; border-radius: 6px; margin: 1rem 0; }}
.controls label {{ font-size: 0.9rem; }}
table {{ border-collapse: collapse; font-size: 0.84rem; margin: 0.75rem 0; background: #fff; }}
th, td {{ border: 1px solid var(--border); padding: 0.4rem 0.55rem; text-align: left; vertical-align: top; }}
th {{ background: #e8ecf1; white-space: nowrap; }}
.same {{ background: #d4edda; }} .changed {{ background: #fff3cd; }}
.escape {{ background: #f8d7da; font-weight: 600; }} .gap {{ background: #e9ecef; color: #666; }}
.gscore {{ color: #555; font-size: 0.8em; font-weight: normal; }}
.mut {{ font-weight: 600; }}
.toggle-conserved {{ margin: 0.35rem 0 0.5rem; padding: 0.35rem 0.75rem; font-size: 0.85rem; cursor: pointer; }}
.treatment-section {{ margin: 1.5rem 0; }}
.treatment-heading {{ display: flex; flex-wrap: wrap; align-items: center; gap: 0.5rem; font-size: 1.15rem; margin: 0 0 0.5rem; border-bottom: 2px solid var(--accent); padding-bottom: 0.35rem; }}
.toggle-treatment {{ margin-left: auto; padding: 0.35rem 0.85rem; font-size: 0.85rem; font-weight: normal; cursor: pointer; border: 1px solid var(--border); background: #fff; border-radius: 6px; }}
.toggle-treatment:hover {{ background: #eef2f7; }}
.treatment-body.treatment-collapsed {{ display: none; }}
.treatment-body {{ padding-top: 0.5rem; }}
.summary-treatment-title {{ font-size: 1rem; margin: 1.1rem 0 0.45rem; color: var(--accent); font-weight: 600; }}
.treatment-body .summary-treatment-title:first-of-type {{ margin-top: 0.25rem; }}
.treatment-body .summary-table + h2 {{ margin-top: 1.35rem; }}
th.sortable {{ cursor: pointer; user-select: none; }}
th.sortable:hover {{ background: #dce3eb; }}
th.sortable.sorted-asc::after {{ content: " ▲"; font-size: 0.75em; }}
th.sortable.sorted-desc::after {{ content: " ▼"; font-size: 0.75em; }}
.epitope-table tr.conserved-hidden {{ display: none; }}
</style></head><body>
<div class="wrap">
<h1>GP mAb escape epitope report</h1>
<p><strong>In silico only</strong> &mdash; not proof of neutralization.</p>

{_render_collapsible_section("intro-therapies", "Therapies in this report", _therapies_intro_html())}

{_render_collapsible_section("intro-grantham", "Grantham distance (numbers in brackets)", _grantham_legend_html())}

<div class="controls">
<label>Min Grantham (hide rows below): <input type="range" id="minGrantham" min="0" max="215" value="0" step="5"> <span id="minGranthamVal">0</span></label>
<label>Sort epitope rows: <select id="sortRows">
<option value="grantham-desc" selected>Max Grantham (high &rarr; low)</option>
<option value="grantham-asc">Max Grantham (low &rarr; high)</option>
<option value="pos">Position (low &rarr; high)</option>
</select></label>
<button type="button" id="applyBtn">Apply to epitope tables</button>
</div>

{tabs_html}
{panels_html}

</div>
<script>
(function() {{
  document.querySelectorAll('.tab-btn').forEach(btn => {{
    btn.addEventListener('click', () => {{
      document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
      document.querySelectorAll('.tab-panel').forEach(p => p.classList.remove('active'));
      btn.classList.add('active');
      const panel = document.getElementById(btn.dataset.panel);
      if (panel) panel.classList.add('active');
    }});
  }});
  const slider = document.getElementById('minGrantham');
  const valSpan = document.getElementById('minGranthamVal');
  const sortSel = document.getElementById('sortRows');
  function updateTableRows(tbl, minG) {{
    const tbody = tbl.querySelector('tbody');
    const rows = Array.from(tbody.querySelectorAll('tr'));
    rows.forEach(tr => {{
      if (tr.classList.contains('conserved-row') && tr.classList.contains('conserved-hidden')) {{
        tr.style.display = 'none';
        return;
      }}
      const rowMax = parseInt(tr.dataset.rowMaxGrantham || '0', 10);
      tr.style.display = rowMax >= minG ? '' : 'none';
    }});
    if (sortSel.value === 'grantham-desc') {{
      rows.sort((a,b) => parseInt(b.dataset.rowMaxGrantham||0,10) - parseInt(a.dataset.rowMaxGrantham||0,10));
    }} else if (sortSel.value === 'grantham-asc') {{
      rows.sort((a,b) => parseInt(a.dataset.rowMaxGrantham||0,10) - parseInt(b.dataset.rowMaxGrantham||0,10));
    }} else {{
      rows.sort((a,b) => parseInt(a.dataset.pos||0,10) - parseInt(b.dataset.pos||0,10));
    }}
    rows.forEach(r => tbody.appendChild(r));
  }}
  function apply() {{
    const minG = parseInt(slider.value, 10) || 0;
    valSpan.textContent = minG;
    const activePanel = document.querySelector('.tab-panel.active') || document.querySelector('.tab-panel');
    if (activePanel) {{
      activePanel.querySelectorAll('table.epitope-table').forEach(tbl => updateTableRows(tbl, minG));
    }}
  }}
  slider.addEventListener('input', apply);
  sortSel.addEventListener('change', apply);
  document.getElementById('applyBtn').addEventListener('click', apply);
  document.querySelectorAll('.toggle-conserved').forEach(btn => {{
    btn.addEventListener('click', () => {{
      const tbl = document.getElementById(btn.dataset.table);
      if (!tbl) return;
      const show = !!tbl.querySelector('tbody tr.conserved-row.conserved-hidden');
      tbl.querySelectorAll('tbody tr.conserved-row').forEach(tr => {{
        if (show) tr.classList.remove('conserved-hidden');
        else tr.classList.add('conserved-hidden');
      }});
      btn.textContent = show ? 'Hide conserved rows' : 'Show conserved rows';
      updateTableRows(tbl, parseInt(slider.value, 10) || 0);
    }});
  }});
  document.querySelectorAll('.toggle-treatment').forEach(btn => {{
    btn.addEventListener('click', () => {{
      const body = document.getElementById(btn.getAttribute('aria-controls'));
      if (!body) return;
      const collapsed = body.classList.toggle('treatment-collapsed');
      btn.setAttribute('aria-expanded', collapsed ? 'false' : 'true');
      const showLabel = btn.getAttribute('data-show-label') || 'Show results';
      const hideLabel = btn.getAttribute('data-hide-label') || 'Hide results';
      btn.textContent = collapsed ? showLabel : hideLabel;
      if (!collapsed) {{
        const panel = body.closest('.tab-panel') || document;
        panel.querySelectorAll('table.epitope-table').forEach(tbl => {{
          updateTableRows(tbl, parseInt(slider.value, 10) || 0);
        }});
      }}
    }});
  }});
  function summaryCellVal(tr, sortKey) {{
    return tr.getAttribute('data-' + sortKey) ?? '';
  }}
  function sortSummaryRows(tbl, sortKey, sortType, asc) {{
    const tbody = tbl.querySelector('tbody');
    if (!tbody) return;
    const rows = Array.from(tbody.querySelectorAll('tr')).filter(tr => summaryCellVal(tr, sortKey) !== '');
    rows.sort((a, b) => {{
      let va = summaryCellVal(a, sortKey);
      let vb = summaryCellVal(b, sortKey);
      if (sortType === 'number') {{
        va = parseFloat(va) || 0;
        vb = parseFloat(vb) || 0;
        return asc ? va - vb : vb - va;
      }}
      va = String(va).toLowerCase();
      vb = String(vb).toLowerCase();
      return asc ? va.localeCompare(vb) : vb.localeCompare(va);
    }});
    rows.forEach(r => tbody.appendChild(r));
  }}
  function markSummarySortHeader(tbl, sortKey, asc) {{
    tbl.querySelectorAll('thead th.sortable').forEach(th => {{
      th.classList.remove('sorted-asc', 'sorted-desc');
      if (th.dataset.sortKey === sortKey) {{
        th.classList.add(asc ? 'sorted-asc' : 'sorted-desc');
      }}
    }});
  }}
  document.querySelectorAll('table.summary-table.sortable-table').forEach(tbl => {{
    const defaultKey = 'conserved';
    sortSummaryRows(tbl, defaultKey, 'number', true);
    markSummarySortHeader(tbl, defaultKey, true);
    tbl.querySelectorAll('thead th.sortable').forEach(th => {{
      th.addEventListener('click', () => {{
        const sortKey = th.dataset.sortKey;
        const sortType = th.dataset.sortType || 'number';
        const asc = !th.classList.contains('sorted-asc');
        sortSummaryRows(tbl, sortKey, sortType, asc);
        markSummarySortHeader(tbl, sortKey, asc);
      }});
    }});
  }});
  apply();
}})();
</script>
</body></html>"""
    with open(path, "w", encoding="utf-8") as f:
        f.write(html)



def _write_escape_verification_files(
    output_dir: str,
    gp_aln_path: str,
    metadata: dict,
    watchlist: List[PositionEntry],
    anchor_id: str,
    pos_to_col: Dict[int, int],
) -> None:
    """Write GP alignment (with Makona row) and coordinate map into Escape/ for cross-checks."""
    from Bio.Seq import Seq
    from Bio.SeqRecord import SeqRecord

    dest_aln = os.path.join(output_dir, "gp_protein_aln.fasta")
    records = list(AlignIO.read(gp_aln_path, "fasta"))
    anchor_row = None
    for rec in records:
        if rec.id == anchor_id:
            anchor_row = str(rec.seq)
            break
    msa_length = len(anchor_row) if anchor_row else len(str(records[0].seq))

    baseline = _baseline_mature_sequence(metadata)
    mature_start = int(metadata.get("mature_gp_start", 33))
    ref_acc = metadata.get("reference_accession", "KJ660346.2")
    makona_header = f"{ref_acc}|Makona_2014_GP_mature|Q05320_baseline"

    makona_ids = {
        r.id for r in records if _is_makona_baseline_row(str(r.seq), baseline, pos_to_col, mature_start)
    }
    makona_in_aln = [r for r in records if r.id in makona_ids]
    other_records = [r for r in records if r.id not in makona_ids]
    out_records: List[SeqRecord] = []

    if makona_in_aln:
        ref_rec = makona_in_aln[0]
        ref_rec.id = makona_header
        ref_rec.description = "Makona 2014 GP in MSA coordinates (report ref_aa / Q05320 baseline)"
        out_records.append(ref_rec)
        out_records.extend(other_records)
        for dup in makona_in_aln[1:]:
            out_records.append(dup)
    else:
        makona_row = _build_makona_msa_row(pos_to_col, baseline, msa_length, mature_start)
        out_records.append(
            SeqRecord(
                Seq(makona_row),
                id=makona_header,
                description="Makona 2014 GP projected into MSA (report ref_aa; not in original alignment)",
            )
        )
        out_records.extend(records)

    SeqIO.write(out_records, dest_aln, "fasta")
    ref_fasta = os.path.join(output_dir, "makona_gp_mature_reference.fasta")
    with open(ref_fasta, "w", encoding="utf-8") as fh:
        fh.write(
            f">{ref_acc} GP mature (Makona 2014; Q05320 numbering baseline for report ref_aa)\n"
        )
        for i in range(0, len(baseline), 80):
            fh.write(baseline[i : i + 80] + "\n")

    map_path = os.path.join(output_dir, "gp_epitope_q05320_msa_map.tsv")
    with open(map_path, "w", encoding="utf-8") as fh:
        fh.write(
            "q05320_pos\tmsa_column_1based\tmsa_anchor_id\tmakona_ref_aa\tdomain\tregion\t"
            "mAb114\tREGN3470\tREGN3471\tREGN3479\tADI15878\tADI23774\tproven_escape\n"
        )
        for ent in sorted(watchlist, key=lambda e: e.pos):
            col = pos_to_col.get(ent.pos)
            col_s = str(col + 1) if col is not None else ""
            esc = ",".join(ent.proven_escape)
            fh.write(
                f"{ent.pos}\t{col_s}\t{anchor_id}\t{ent.zaire_aa}\t{ent.domain}\t{ent.region}\t"
                f"{int(ent.mAb114)}\t{int(ent.REGN3470)}\t{int(ent.REGN3471)}\t{int(ent.REGN3479)}\t"
                f"{int(ent.ADI15878)}\t{int(ent.ADI23774)}\t{esc}\n"
            )

    print(
        f"mab-escape-report: wrote {dest_aln} ({len(out_records)} sequences, "
        f"Makona baseline {'in original MSA' if makona_in_aln else 'added as first row'})"
    )
    print(f"mab-escape-report: wrote {ref_fasta}")
    print(f"mab-escape-report: wrote {map_path}")


def run_mab_escape_report(gp_aln_path: str, output_dir: str = "Escape", watchlist_path: Optional[str] = None):
    os.makedirs(output_dir, exist_ok=True)
    metadata, watchlist = load_watchlist(watchlist_path)
    ref_id, pos_to_col, matrix, summaries, excluded = score_alignment(gp_aln_path, watchlist, metadata)
    _write_escape_verification_files(output_dir, gp_aln_path, metadata, watchlist, ref_id, pos_to_col)
    html_path = os.path.join(output_dir, "gp_mab_escape_report.html")
    write_html(html_path, gp_aln_path, metadata, watchlist, ref_id, matrix, summaries, excluded)
    xlsx_path = write_r_workbook(
        output_dir, gp_aln_path, watchlist, ref_id, matrix, summaries, metadata
    )
    print(f"mab-escape-report: wrote {html_path}")
    print(f"mab-escape-report: wrote {xlsx_path}")
    return html_path
